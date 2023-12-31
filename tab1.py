# tab1.py
# 식품안전나라 데이터 불러오기 페이지
# pip install requests matplotlib openpyxl numpy

import streamlit as st
import requests
import pprint
import json
import smtplib
import datetime
import pandas as pd
import re
import time
from openpyxl import load_workbook, Workbook
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import os
import zipfile





pro = ["lactobacillus", "bifidobacterium", "synbiotics", "prebiotics", "probiotics", "postbiotics", 
    "parabiotics", "프로바이오틱스", "프리바이오틱스", "신바이오틱스", "포스트바이오틱스", "acidophilus", "casei", 
           "gasseri", "delbrueck", "bulgaricus", "helveticus", "fermentum", "paracasei", "바이오","틱스",
           "plantarum", "reuteri", "rhamnosus", "salivarius", "lactis", "faecium", "락토","비피도",
           "faecalis", "thermophilus", "bifidum", "breve", "longum", "animalis","lactis", 
       "락토바실러스", "비피도박테리움","비피더스", "스트렙토코커스",
       'Lactobacillus', 'Bifidobacterium', 'Synbiotics', 'Prebiotics', 'Probiotics', 'Postbiotics', 'Parabiotics',
       'Acidophilus', 'Casei', 'Gasseri', 'Delbrueck', 'Bulgaricus', 'Helveticus', 'Fermentum', 
       'Paracasei', 'Plantarum', 'Reuteri', 'Rhamnosus', 'Salivarius', 'Lactis', 'Faecium', 
    'Faecalis', 'Thermophilus', 'Bifidum', 'Breve', 'Longum', 'Animalis', 'Lactis']

vita = ["비타민 A", "비타민A", "베타카로틴", "비타민 D", "비타민D", "비타민 E", "비타민E", 
    "비타민 K", "비타민K", "비타민 B1", "비타민B1", "비타민 B2", "B2", "나이아신","판토텐산",
    "비타민 B6", "B6", "엽산", "비타민 B12", "B12", "비오틴", "비타민 C", "비타민C","칼슘",
    "마그네슘", "마그늄", "철","아연","구리","셀레늄", "셀렌","요오드","망간", "망가니즈",
        "몰리브덴","칼륨","크롬","필수지방산","단백질","식이섬유"]

vita3 = ["비타민 A", "베타카로틴", "비타민 D", "비타민 E", "비타민 K", "비타민 B1", "비타민 B2",
         "나이아신","판토텐산", "비타민 B6", "엽산", "비타민 B12", "비오틴", "비타민 C", "칼슘",
    "마그네슘", "철","아연","구리","셀레늄", "셀렌","요오드","망간", 
        "몰리브덴","칼륨","크롬","필수지방산","단백질","식이섬유", "셀레늄(또는 셀렌)"]

vita2 = ["A", "B", "B1", "B2", "B12", 'D','E', 'B6', 'C']

gosi = ['인삼', '홍삼', '엽록소', '클로렐라','스피루리나','녹차', '알로에전잎','알로에 전잎', '프로폴리스', 
        '코엔자임 q10','코엔자임q10','코엔자임 Q10','코엔자임Q10', 'Q10',
        '코엔자임', '대두이소플라본', '구아바잎','구아바 잎', '바나바잎', '바나바 잎', '은행잎',
        '은행 잎', '카르두스마리아누스', '밀크씨슬', '달맞이꽃','달맞이 꽃', 'DHA', 'EPA', '감마리놀', '레시틴', 
        '스쿠알렌', '식물스테롤', '알콕시글리세롤',
        '상어간유', '옥타코사놀', '매실', '공액리놀레산', '가르시니아', '루테인', '헤마토', '쏘팔메토',
        '포스파티딜', '글루코사민', 'N-아세틸글루코사민', 'N-아세틸', 'NAG',
        'N 아세틸글루코사민', 'N 아세틸', 
        'n-아세틸글루코사민', 'n아세틸글루코사민','n 아세틸글루코사민', 
        'n-아세틸','n아세틸','n 아세틸', 'nag', '뮤코다당․단백', '뮤코다당',
        '알로에겔','알로에 겔', '영지버섯', '키토올리고당', '키토산', '키토올리고당', '프락토올리고당', '프로바이오틱스', '홍국',
        '대두단백', '테아닌', '엠에스엠', 'MSM',  '폴리감마글루', '마늘', '히알루론산', '홍경천', '빌베리', '라피노스',
        '크레아틴', '유단백가수', '상황버섯', '토마토', '곤약감자', '구아검','글루코만난', '곤약', '귀리식이', '난소화성말토덱스트린',
        '난소화성말토', '대두식이', '목이버섯식이', '밀식이', '보리식이', '아카시아검', '아라비아검', '옥수수겨식이',
        '치커리추출물', '이눌린', '치커리', '차전자피', '폴리덱스트로', '호로파종자', '분말한천', '화화나무'] 


epa = ["오메가-3", "오메가", "오메가 3", "오메가3"]


page3_item = ["멀티비타민미네랄", "다이어트", "장건강","뼈", "관절", "항산화", "혈행", "혈당", "면역",
              "mind 건강", "Brain 건강", "눈", "간", "배변활동", "이너뷰티", "여성건강", "남성건강", "단백질"]

page4_item = ["프로바이오틱스", "비타민미네랄","가르시니아","EPA및DHA","홍삼","밀크씨슬","칼마디","MSM / NAG","비타민C",
              "비오틴","비타민BC","눈 건강","프로폴리스","차전자피식이섬유","쏘팔메토/옥타코사놀",
              "바나바잎추출","은행잎추출","콜라겐", "크레아틴", "포스파티딜세린", "감마리놀렌산"]

kmd = ["칼슘", "마그네슘", "비타민D", "비타민 D"]
# 파일 받기
def received_contents(request_url):
    response = requests.get(request_url)
    contents = response.text
    pp = pprint.PrettyPrinter(indent=4)
    return contents

# 받은 텍스트 json으로 변환
def translation_json(contents):
    json_ob = json.loads(contents)
    return json_ob

def week_of_month():
    return_li = []
    today = datetime.date.today()    
    month = int(today.strftime("%m"))
    return_li.append(month)
    
    first_day_month = today.replace(day=1)
    day_of_week = first_day_month.weekday()
    day_of_month = today.day
    week_of_month = (day_of_month + day_of_week) // 7 + 1
    return_li.append(week_of_month)

    return return_li

def load_while(load_ws, min_date, max_date):
    progress_text = "Download in progress. Please wait."
    my_bar = st.progress(0, text=progress_text)
    
    cnt = 1
    key = "8a6b626961064df497a3"
    request_start = 1
    request_finish = 1000
    request_url = "http://openapi.foodsafetykorea.go.kr/api/%20s/I0030/json/%d/%d" % (key, request_start,request_finish)

    min_date = min_date.strftime("%Y%m%d")
    max_date = max_date.strftime("%Y%m%d")
    whe = 1
    while True:
        my_bar.progress(int(cnt * (100/40)) , text=progress_text) ## /40 변수는 업데이트 해야함.
        contents = received_contents(request_url)
        json_ob = translation_json(contents)
        
        
        for i in range(int(json_ob["I0030"]["total_count"])):
            # print("=>for start i: ",i)

            PRMS_DT = json_ob["I0030"]["row"][i]["PRMS_DT"] # PRMS_DT -> 허가_일자 5
            if(min_date <= PRMS_DT <= max_date):
                    
                LCNS_NO = json_ob["I0030"]["row"][i]["LCNS_NO"] # LCNS_NO -> 인허가번호 1
                BSSH_NM = json_ob["I0030"]["row"][i]["BSSH_NM"] # BSSH_NM -> 업소_명 2
                PRDLST_REPORT_NO = json_ob["I0030"]["row"][i]["PRDLST_REPORT_NO"] # PRDLST_REPORT_NO -> 품목제조번호 3
                PRDLST_NM = json_ob["I0030"]["row"][i]["PRDLST_NM"] # PRDLST_NM -> 품목_명 4
                POG_DAYCNT = json_ob["I0030"]["row"][i]["POG_DAYCNT"] # POG_DAYCNT -> 유통/소비기한_일수 6
                DISPOS = json_ob["I0030"]["row"][i]["DISPOS"] # DISPOS -> 성상 7
                NTK_MTHD = json_ob["I0030"]["row"][i]["NTK_MTHD"] # NTK_MTHD -> 섭취방법 8
                PRIMARY_FNCLTY = json_ob["I0030"]["row"][i]["PRIMARY_FNCLTY"] # PRIMARY_FNCLTY -> 주된기능성 9
                IFTKN_ATNT_MATR_CN = json_ob["I0030"]["row"][i]["IFTKN_ATNT_MATR_CN"] # IFTKN_ATNT_MATR_CN -> 섭취시주의사항 10
                
                #####

                CSTDY_MTHD = json_ob["I0030"]["row"][i]["CSTDY_MTHD"] # CSTDY_MTHD -> 보관방법 11
                PRDLST_CDNM = json_ob["I0030"]["row"][i]["PRDLST_CDNM"] # PRDLST_CDNM -> 유형 12
                STDR_STND = json_ob["I0030"]["row"][i]["STDR_STND"] # STDR_STND -> 기준규격 13
                HIENG_LNTRT_DVS_NM = json_ob["I0030"]["row"][i]["HIENG_LNTRT_DVS_NM"] # HIENG_LNTRT_DVS_NM -> 고열량저영양여부 14
                PRODUCTION = json_ob["I0030"]["row"][i]["PRODUCTION"] # PRODUCTION -> 생산종료여부 15 
                CHILD_CRTFC_YN = json_ob["I0030"]["row"][i]["CHILD_CRTFC_YN"] # CHILD_CRTFC_YN -> 어린이기호식품품질인증여부 16
                PRDT_SHAP_CD_NM = json_ob["I0030"]["row"][i]["PRDT_SHAP_CD_NM"] # PRDT_SHAP_CD_NM -> 제품형태 17
                FRMLC_MTRQLT = json_ob["I0030"]["row"][i]["FRMLC_MTRQLT"] # FRMLC_MTRQLT -> 포장재질 18
                RAWMTRL_NM = json_ob["I0030"]["row"][i]["RAWMTRL_NM"] # RAWMTRL_NM -> 품목유형(기능지표성분) 19
                INDUTY_CD_NM = json_ob["I0030"]["row"][i]["INDUTY_CD_NM"] # INDUTY_CD_NM -> 업종 20

                #####

                LAST_UPDT_DTM = json_ob["I0030"]["row"][i]["LAST_UPDT_DTM"] # LAST_UPDT_DTM -> 최종수정일자 21
                INDIV_RAWMTRL_NM = json_ob["I0030"]["row"][i]["INDIV_RAWMTRL_NM"] # INDIV_RAWMTRL_NM -> 기능성 원재료 22
                ETC_RAWMTRL_NM = json_ob["I0030"]["row"][i]["ETC_RAWMTRL_NM"] # ETC_RAWMTRL_NM -> 기타 원재료 23
                CAP_RAWMTRL_NM = json_ob["I0030"]["row"][i]["CAP_RAWMTRL_NM"] # CAP_RAWMTRL_NM -> 캡슐 원재료 24

                #####

                # 행&열 작업
                s1 = "A"+ str(whe + 1)
                s2 = "B"+ str(whe + 1)
                s3 = "C"+ str(whe + 1)
                s4 = "D"+ str(whe + 1)
                s5 = "E"+ str(whe + 1)
                s6 = "F"+ str(whe + 1)
                s7 = "G"+ str(whe + 1)
                s8 = "H"+ str(whe + 1)
                s9 = "I"+ str(whe + 1)
                s10 = "J"+ str(whe + 1)
                s11 = "K"+ str(whe + 1)
                s12 = "L"+ str(whe + 1)
                s13 = "M"+ str(whe + 1)
                s14 = "N"+ str(whe + 1)
                s15 = "O"+ str(whe + 1)
                s16 = "P"+ str(whe + 1)
                s17 = "Q"+ str(whe + 1)
                s18 = "R"+ str(whe + 1)
                
                load_ws[s1] = PRMS_DT
                load_ws[s2] = LAST_UPDT_DTM
                load_ws[s3] = PRDT_SHAP_CD_NM
                load_ws[s4] = PRDLST_NM
                load_ws[s5] = STDR_STND
                load_ws[s6] = BSSH_NM
                load_ws[s7] = RAWMTRL_NM
                load_ws[s8] = INDIV_RAWMTRL_NM
                load_ws[s9] = ETC_RAWMTRL_NM
                load_ws[s10] = DISPOS
                load_ws[s11] = NTK_MTHD
                load_ws[s12] = PRIMARY_FNCLTY
                load_ws[s13] = FRMLC_MTRQLT
                load_ws[s14] = POG_DAYCNT
                load_ws[s15] = CAP_RAWMTRL_NM
                load_ws[s16] = IFTKN_ATNT_MATR_CN
                load_ws[s17] = PRDLST_REPORT_NO
                load_ws[s18] = LCNS_NO
                whe += 1
            else:
                continue
        if(int(json_ob["I0030"]["total_count"]) != 1000):
            print("총 갯수:", cnt*1000 + int(json_ob["I0030"]["total_count"]))
            my_bar.progress(100 , text="Download process finish") ## /40 변수는 업데이트 해야함.
            break

        request_start += 1000
        request_finish += 1000
        request_url = "http://openapi.foodsafetykorea.go.kr/api/%20s/I0030/json/%d/%d" % (key, request_start,request_finish)
        cnt += 1

        # 테스트용
        # if(cnt == 3):
        #     my_bar.progress(100 , text="Download process finish") ## /40 변수는 업데이트 해야함.
        #     break
        
def page2(df, file_name, load_ws2):
    page2_tot = [[[i,0] for i in range(13)]for j in range(5)]
    page2_uni = [[[i,0] for i in range(13)]for j in range(5)]
    page2_exp = [[[i,0] for i in range(13)]for j in range(5)]

    for i in range(0,len(df)):
        ss = df_until_now.iloc[i]
        page2_tot[2023 - int(ss["year"])][ss["month"]-1][1] += 1
        try:
            if(str(ss["기능성"]).count("(제20") >=1 ):
                page2_uni[2023 - int(ss["year"])][ss["month"]-1][1] += 1
            if(str(ss["제품명"]).count("수출") >=1):
                page2_exp[2023 - int(ss["year"])][ss["month"]-1][1] += 1
        except:
            continue

    

    # 행&열 작업
    line_B = [["B" for i in range(13)] for j in range(5)]
    line_C = [["C" for i in range(13)] for j in range(5)]
    line_D = [["D" for i in range(13)] for j in range(5)]
    line_E = [["E" for i in range(13)] for j in range(5)]
    
    
    for j in range(5):
        for i in range(13):
            line_B[j][i] += str(i+5 +(19*j))
            line_C[j][i] += str(i+5 +(19*j))
            line_D[j][i] += str(i+5 +(19*j))
            line_E[j][i] += str(i+5 +(19*j))
#     print(page2_tot)
    for j in range(5):
        for i in range(13):
            if(i == 12):
                sum_values = 0
                for data in page2_tot[j]:
                    sum_values += data[1]
                load_ws2[line_B[j][i]] = sum_values
                
                sum_values = 0
                for data in page2_uni[j]:
                    sum_values += data[1]
                load_ws2[line_C[j][i]] = sum_values
                
                sum_values = 0
                for data in page2_exp[j]:
                    sum_values += data[1]
                load_ws2[line_E[j][i]] = sum_values
                
            else:
                if(page2_tot[j][i][1] != 0):
                    load_ws2[line_B[j][i]] = page2_tot[j][i][1]
                    load_ws2[line_C[j][i]] = page2_uni[j][i][1]
                    load_ws2[line_D[j][i]] = page2_uni[j][i][1]/page2_tot[j][i][1]*100
                    load_ws2[line_E[j][i]] = int(page2_exp[j][i][1])

def is_list_in_set(lst, s):
    # 리스트의 값들이 모두 집합에 포함되어 있고, 집합의 크기가 리스트의 크기보다 크거나 같으면 True, 그렇지 않으면 False 반환
    ori_s = len(s)
    if(ori_s > 0 ):
        for i in lst:
            s.discard(i)
        return [len(s), ori_s]
    else:
        return [-1, -1]
def page3(df, file_name, load_ws3):
    page3_ori = [[[["기초영양소",0], ["다이어트", 0], ["장건강", 0], ["뼈", 0], ["관절", 0], ["항산화", 0], 
                  ["혈행", 0], ["혈당", 0], ["면역", 0], ["mind 건강", 0], ["Brain 건강", 0], ["눈", 0], ["간", 0],
                  ["배변활동", 0], ["이너뷰티", 0], ["여성건강", 0], ["남성건강", 0], ["단백질", 0]] for i in range(13)]for j in range(5)]
        
    for i in range(len(df)):
        row = df.iloc[i]
        r_year = int(row['year'])
        r_month = int(row['month'])
        r_gi = str(row['기능성'])
        r_gin = str(row['기능성내용'])                
        r_ju = str(row['주원료'])
        my_set = set(r_gi.split(", "))

        
        ##### 기초영양소
        flag1 = False
        flag2 = False
        for j in gosi: # 고시형 제외
            if(r_gi.count(j)>=1):
                flag1 = True
                break

        for j in vita: # 비타민 미네랄 포함
            if(r_gi.count(j)>=1):
                flag2 = True
                break

        if( not flag1 and flag2 and r_gi.count("제20")==0):
#             print("r_gi:", r_gi)
            page3_ori[2023 - r_year][r_month-1][0][1]+=1


        ##### 다이어트    
        if(r_gin.count("체지방") >= 1):
            page3_ori[2023 - r_year][r_month-1][1][1]+=1

        ##### 장건강
        if((r_gin.count("장 건강") >= 1 or r_gin.count("장건강") >= 1) and (str(row["제품명"]).count("혼합") == 0)):
            if (any(keyword in r_gi for keyword in pro) or
            r_gi.count("알로에 겔") >= 1):
                page3_ori[2023 - r_year][r_month-1][2][1]+=1 

        ##### 뼈
        if(r_gin.count("뼈 건강") >= 1 or r_gin.count("뼈건강") >= 1 or 
          r_gin.count("뼈")>=1):
            if(is_list_in_set(['칼슘', '망간', 'D','비타민 D','비타민D', 'K', '비타민K','비타민 K', '마그네슘', '폴리감마글루탐산'],my_set)[0] == 0): 
                page3_ori[2023 - r_year][r_month-1][3][1]+=1
                
        my_set = set(r_gi.split(", "))

        ##### 관절
        if(r_gin.count("관절") >= 1):
            page3_ori[2023 - r_year][r_month-1][4][1]+=1

        ##### 항산화
        if(r_gin.count("항산화") >= 1 and (r_gi.count("코엔자임") >= 1 or 
           r_gi.count("프로폴리스") >= 1)):
            page3_ori[2023 - r_year][r_month-1][5][1]+=1

        ##### 혈행
        if(r_gin.count("혈행") >= 1 and r_gi.count("EPA") >= 1):
            page3_ori[2023 - r_year][r_month-1][6][1]+=1

        ##### 혈당
        if(r_gin.count("혈당") >= 1):
            page3_ori[2023 - r_year][r_month-1][7][1]+=1

        ##### 면역력
        if((r_gin.count("면역력") >= 1 or r_gin.count("면역기능") >= 1) and 
          ((r_gi.count("아연") >= 1 or r_gi.count("베타글로칸") >= 1 or
           r_gi.count("알로에") >= 1) and r_gi.count("홍삼") == 0)):
            page3_ori[2023 - r_year][r_month-1][8][1]+=1

        ##### mind 건강
        if(r_gin.count("스트레스") >= 1 or r_gin.count("수면") >= 1 or 
          r_gin.count("긴장") >= 1 or r_gin.count("피로") >= 1):
            page3_ori[2023 - r_year][r_month-1][9][1]+=1

        ##### brain 건강
        if(r_gin.count("인지") >= 1 or r_gin.count("기억") >= 1):
            page3_ori[2023 - r_year][r_month-1][10][1]+=1

        ##### 눈 건강
        if(r_gin.count("눈") >= 1 and 
           (any(keyword in r_gi for keyword in epa) or r_gi.count("마리골드") >= 1 or 
            r_gi.count("루테인") >= 1)):
            page3_ori[2023 - r_year][r_month-1][11][1]+=1

        ##### 간
        if(r_gin.count("간 건강") >= 1 or r_gin.count("간건강") >= 1):
            page3_ori[2023 - r_year][r_month-1][12][1]+=1

        ##### 배변
        if(r_gin.count("배변활동") >= 1 or r_gin.count("배변 활동") >= 1):
            page3_ori[2023 - r_year][r_month-1][13][1]+=1

        if(r_gin.count("보습") >= 1 or r_gin.count("자외선") >= 1):
            page3_ori[2023 - r_year][r_month-1][14][1]+=1

        if(r_gin.count("여성") >= 1 and r_gi.count("홍삼") == 0):
            page3_ori[2023 - r_year][r_month-1][15][1]+=1

        if(r_gin.count("남성") >= 1 or r_gin.count("전립선") >= 1):
            page3_ori[2023 - r_year][r_month-1][16][1]+=1
            
        if(r_gi.count("단백질") >= 1):
            page3_ori[2023 - r_year][r_month-1][17][1]+=1
    
    # 행&열 작업
    line_G = [["G" for i in range(20)] for j in range(5)]
    line_H = [["H" for i in range(20)] for j in range(5)]
    line_I = [["I" for i in range(20)] for j in range(5)]
    line_J = [["J" for i in range(20)] for j in range(5)]
    line_K = [["K" for i in range(20)] for j in range(5)]
    line_L = [["L" for i in range(20)] for j in range(5)]
    line_M = [["M" for i in range(20)] for j in range(5)]
    line_N = [["N" for i in range(20)] for j in range(5)]
    line_O = [["O" for i in range(20)] for j in range(5)]
    line_P = [["P" for i in range(20)] for j in range(5)]
    line_Q = [["Q" for i in range(20)] for j in range(5)]
    line_R = [["R" for i in range(20)] for j in range(5)]
    line_S = [["S" for i in range(20)] for j in range(5)]

    for j in range(5):
        for i in range(20):
            line_G[j][i] += str(i+4 + (25*j))
            line_H[j][i] += str(i+4 + (25*j))
            line_I[j][i] += str(i+4 + (25*j))
            line_J[j][i] += str(i+4 + (25*j))
            line_K[j][i] += str(i+4 + (25*j))
            line_L[j][i] += str(i+4 + (25*j))
            line_M[j][i] += str(i+4 + (25*j))
            line_N[j][i] += str(i+4 + (25*j))
            line_O[j][i] += str(i+4 + (25*j))
            line_P[j][i] += str(i+4 + (25*j))
            line_Q[j][i] += str(i+4 + (25*j))
            line_R[j][i] += str(i+4 + (25*j))
            line_S[j][i] += str(i+4 + (25*j))

    for k in range(5):
        for i in range(13):
            for j in range(19):
                if(j == 18):
                    j+=1
                    count = 0
                    # 행부터 채우기
                    for data in page3_ori[k]:
                        count += 1
                        sum_values = 0
                        for sublist in data: 
                            sum_values += int(sublist[1])
                        if(count == 1):
                            load_ws3[line_G[k][j]] = sum_values
                        elif(count == 2):
                            load_ws3[line_H[k][j]] = sum_values
                        elif(count == 3):
                            load_ws3[line_I[k][j]] = sum_values
                        elif(count == 4):
                            load_ws3[line_J[k][j]] = sum_values
                        elif(count == 5):
                            load_ws3[line_K[k][j]] = sum_values
                        elif(count == 6):
                            load_ws3[line_L[k][j]] = sum_values
                        elif(count == 7):
                            load_ws3[line_M[k][j]] = sum_values
                        elif(count == 8):
                            load_ws3[line_N[k][j]] = sum_values
                        elif(count == 9):
                            load_ws3[line_O[k][j]] = sum_values
                        elif(count == 10):
                            load_ws3[line_P[k][j]] = sum_values
                        elif(count == 11):
                            load_ws3[line_Q[k][j]] = sum_values
                        elif(count == 12):
                            load_ws3[line_R[k][j]] = sum_values
                else:
                    if(i==0):
                        load_ws3[line_G[k][j]] = page3_ori[k][i][j][1]
                        continue
                    elif(i==1):
                        load_ws3[line_H[k][j]] = page3_ori[k][i][j][1]
                        continue
                    elif(i==2):
                        load_ws3[line_I[k][j]] = page3_ori[k][i][j][1]
                        continue
                    elif(i==3):
                        load_ws3[line_J[k][j]] = page3_ori[k][i][j][1]
                        continue
                    elif(i==4):
                        load_ws3[line_K[k][j]] = page3_ori[k][i][j][1]
                        continue
                    elif(i==5):
                        load_ws3[line_L[k][j]] = page3_ori[k][i][j][1]
                        continue
                    elif(i==6):
                        load_ws3[line_M[k][j]] = page3_ori[k][i][j][1]
                        continue
                    elif(i==7):
                        load_ws3[line_N[k][j]] = page3_ori[k][i][j][1]
                        continue
                    elif(i==8):
                        load_ws3[line_O[k][j]] = page3_ori[k][i][j][1]
                        continue
                    elif(i==9):
                        load_ws3[line_P[k][j]] = page3_ori[k][i][j][1]
                        continue
                    elif(i==10):
                        load_ws3[line_Q[k][j]] = page3_ori[k][i][j][1]
                        continue
                    elif(i==11):
                        load_ws3[line_R[k][j]] = page3_ori[k][i][j][1]
                        continue
            
        for o in range(18):
            count_t = 0
            for p in range(12):
                count_t += page3_ori[k][p][o][1]
            load_ws3[line_S[k][o]] = count_t

    return page3_ori




def page4(df, file_name, load_ws4): 
    page4_ori = [[[["프로바이오틱스",0], ["비타민미네랄", 0], ["가르시니아", 0], ["EPA및DHA", 0], 
              ["홍삼", 0], ["밀크씨슬", 0], ["칼마디", 0], ["MSM / NAG", 0], ["비타민C", 0], 
              ["비오틴", 0], ["L-테아닌", 0], ["눈 건강", 0], ["프로폴리스", 0], 
              ["차전자피식이섬유", 0], ["쏘팔메토/옥타코사놀", 0], ["바나바잎추출", 0], 
              ["은행잎추출", 0], ["콜라겐", 0], ["크레아틴", 0], ["포스파티딜세린", 0], 
                  ["감마리놀렌산", 0]] for i in range(13)]for j in range(5)]

   
    for i in range(len(df)):
        row = df.iloc[i]
        r_year = int(row['year'])
        r_month = int(row['month'])
        r_gi = str(row['기능성'])
        r_gin = str(row['기능성내용'])
        r_ju = str(row['주원료'])
        my_set_gi = set(r_gi.split(", "))
        my_set_ju = set(r_ju.split(", "))

        ##### 프로바이오틱스        
        if((r_gi.count("프로바이오틱스") >= 1) and 
           (str(df.loc[i]["제품명"]).count("혼합") == 0 ) and 
           (str(df.loc[i]["제품명"]).count("분말") == 0)):
            page4_ori[2023 - r_year][r_month-1][0][1]+=1
            
         #### 비타민 미네랄 항목
        flag1 = False
        for j in gosi:
            if(r_ju.count(j)>=1):
                flag1 = True
                break

        for j in vita:
            if(r_ju.count(j)>=1):
                if( flag1 == False and r_gi.count("제20")==0):
                    page4_ori[2023 - r_year][r_month-1][1][1]+=1
            break
        
        
        ##### 가르시니아
        if(r_ju.count("가르시니아") >= 1 ):
            page4_ori[2023 - r_year][r_month-1][2][1]+=1

        ##### EPA & DHA
        if(r_gi.count("EPA") >= 1): 
            page4_ori[2023 - r_year][r_month-1][3][1]+=1

        ##### 홍삼
        if(r_ju.count("홍삼") >= 1 ):
            page4_ori[2023 - r_year][r_month-1][4][1]+=1

        ##### 밀크씨슬
        if(r_ju.count("밀크씨슬") >= 1 ):
            page4_ori[2023 - r_year][r_month-1][5][1]+=1

        ##### 칼마디
        if(r_ju.count("캄슘") >= 1 or r_ju.count("마그네슘") >= 1 or r_ju.count("비타민 D") >= 1 or r_ju.count("비타민D") >= 1):
            page4_ori[2023 - r_year][r_month-1][6][1]+=1
        
        ##### n-아세틸
        if(r_ju.count("엠에스엠") >= 1 or r_ju.count("N-아세틸") >= 1 or
          r_ju.count("MSM") >= 1 or r_ju.count("NAG") >= 1 or
          r_ju.count("N - 아세틸") >= 1): ##
            page4_ori[2023 - r_year][r_month-1][7][1]+=1
        my_set = set(r_gi.split(", "))
        
        ##### 비타민 C
        if(is_list_in_set(["비타민C", "비타민 C", "C"],my_set_gi)[0] == 0): 
            page4_ori[2023 - r_year][r_month-1][8][1]+=1
        
        ##### 비오틴
        if(r_ju.count("비오틴") >= 1 or r_ju.count("판토텐산") >= 1):
            page4_ori[2023 - r_year][r_month-1][9][1]+=1


        #### L-테아닌
        if(r_gi.count("테아닌") >= 1):
            page4_ori[2023 - r_year][r_month-1][10][1]+=1

        ##### 눈 건강
        if(r_ju.count("마리골드") >= 1 or r_ju.count("지아잔틴") >= 1): ##
            page4_ori[2023 - r_year][r_month-1][11][1]+=1

        ##### 프로폴리스
        if(r_ju.count("프로폴리스") >= 1 ):
            page4_ori[2023 - r_year][r_month-1][12][1]+=1

        ##### 차전자피
        if(r_ju.count("차전자피") >= 1 ):
            page4_ori[2023 - r_year][r_month-1][13][1]+=1

        ##### 쏘팔메토
        if(r_ju.count("쏘팔메토") >= 1 or r_ju.count("옥타코사놀") >= 1): ##
            page4_ori[2023 - r_year][r_month-1][14][1]+=1

        ##### 바나나
        if(r_ju.count("바나바") >= 1 ):
            page4_ori[2023 - r_year][r_month-1][15][1]+=1

        ##### 은행잎
        if(r_ju.count("은행잎") >= 1 ):
            page4_ori[2023 - r_year][r_month-1][16][1]+=1

        ##### 콜라겐
        if(r_ju.count("콜라겐") >= 1 ):
            page4_ori[2023 - r_year][r_month-1][17][1]+=1

        ##### 크레아틴
        if(r_ju.count("크레아틴") >= 1 ):
            page4_ori[2023 - r_year][r_month-1][18][1]+=1

        ##### 포스파티딜세린
        if(r_ju.count("포스파티딜세린") >= 1 ):
            page4_ori[2023 - r_year][r_month-1][19][1]+=1

        ##### 감마리놀렌산
        if(r_gi.count("감마리놀렌산") >= 1 ):
            page4_ori[2023 - r_year][r_month-1][20][1]+=1



    # 행&열 작업
    line_F = [["F" for i in range(23)] for i in range(5)]
    line_G = [["G" for i in range(23)] for j in range(5)]
    line_H = [["H" for i in range(23)] for j in range(5)]
    line_I = [["I" for i in range(23)] for j in range(5)]
    line_J = [["J" for i in range(23)] for j in range(5)]
    line_K = [["K" for i in range(23)] for j in range(5)]
    line_L = [["L" for i in range(23)] for j in range(5)]
    line_M = [["M" for i in range(23)] for j in range(5)]
    line_N = [["N" for i in range(23)] for j in range(5)]
    line_O = [["O" for i in range(23)] for j in range(5)]
    line_P = [["P" for i in range(23)] for j in range(5)]
    line_Q = [["Q" for i in range(23)] for j in range(5)]
    line_R = [["R" for i in range(23)] for j in range(5)]


    for j in range(5):
        for i in range(23):
            line_F[j][i] += str(i+5 + (27*j))
            line_G[j][i] += str(i+5 + (27*j))
            line_H[j][i] += str(i+5 + (27*j))
            line_I[j][i] += str(i+5 + (27*j))
            line_J[j][i] += str(i+5 + (27*j))
            line_K[j][i] += str(i+5 + (27*j))
            line_L[j][i] += str(i+5 + (27*j))
            line_M[j][i] += str(i+5 + (27*j))
            line_N[j][i] += str(i+5 + (27*j))
            line_O[j][i] += str(i+5 + (27*j))
            line_P[j][i] += str(i+5 + (27*j))
            line_Q[j][i] += str(i+5 + (27*j))
            line_R[j][i] += str(i+5 + (27*j))

    
    
    for k in range(5):
        for i in range(13):
            for j in range(22):
                if(j == 21):
                    j += 1
                    count = 0
                    # 행부터 채우기
                    for data in page4_ori[k]:
                        count += 1
                        sum_values = 0
                        for sublist in data: 
                            sum_values += int(sublist[1])
                        if(count == 1):
                            load_ws4[line_F[k][j]] = sum_values 
                        elif(count == 2):
                            load_ws4[line_G[k][j]] = sum_values
                        elif(count == 3):
                            load_ws4[line_H[k][j]] = sum_values
                        elif(count == 4):
                            load_ws4[line_I[k][j]] = sum_values
                        elif(count == 5):
                            load_ws4[line_J[k][j]] = sum_values
                        elif(count == 6):
                            load_ws4[line_K[k][j]] = sum_values
                        elif(count == 7):
                            load_ws4[line_L[k][j]] = sum_values
                        elif(count == 8):
                            load_ws4[line_M[k][j]] = sum_values
                        elif(count == 9):
                            load_ws4[line_N[k][j]] = sum_values
                        elif(count == 10):
                            load_ws4[line_O[k][j]] = sum_values
                        elif(count == 11):
                            load_ws4[line_P[k][j]] = sum_values
                        elif(count == 12):
                            load_ws4[line_Q[k][j]] = sum_values
                else:
                    if(i==0):
                        load_ws4[line_F[k][j]] = page4_ori[k][i][j][1]
                        continue
                    elif(i==1):
                        load_ws4[line_G[k][j]] = page4_ori[k][i][j][1]
                        continue
                    elif(i==2):
                        load_ws4[line_H[k][j]] = page4_ori[k][i][j][1]
                        continue
                    elif(i==3):
                        load_ws4[line_I[k][j]] = page4_ori[k][i][j][1]
                        continue
                    elif(i==4):
                        load_ws4[line_J[k][j]] = page4_ori[k][i][j][1]
                        continue
                    elif(i==5):
                        load_ws4[line_K[k][j]] = page4_ori[k][i][j][1]
                        continue
                    elif(i==6):
                        load_ws4[line_L[k][j]] = page4_ori[k][i][j][1]
                        continue
                    elif(i==7):
                        load_ws4[line_M[k][j]] = page4_ori[k][i][j][1]
                        continue
                    elif(i==8):
                        load_ws4[line_N[k][j]] = page4_ori[k][i][j][1]
                        continue
                    elif(i==9):
                        load_ws4[line_O[k][j]] = page4_ori[k][i][j][1]
                        continue
                    elif(i==10):
                        load_ws4[line_P[k][j]] = page4_ori[k][i][j][1]
                        continue
                    elif(i==11):
                        load_ws4[line_Q[k][j]] = page4_ori[k][i][j][1]
                        continue
                        
        for o in range(21):
            count_t = 0
            for p in range(12):
                count_t += page4_ori[k][p][o][1]
            load_ws4[line_R[k][o]] = count_t

    return page4_ori



# 고모한테 자동으로 이메일 보내기
def send(file_name, email):
    # smpt 서버와 연결
    gmail_smtp = "smtp.gmail.com"  #gmail smtp 주소
    gmail_port = 465  #gmail smtp 포트번호
    smpt = smtplib.SMTP_SSL(gmail_smtp, gmail_port)

    # 로그인
    my_id = "no1gunhee@gmail.com"
    my_password = "jxopsxbhhzgsaiyc"
    smpt.login(my_id, my_password)


    # 메일 기본 정보 설정
    msg = MIMEMultipart()
    msg["Subject"] = f"건강기능식품 품목 신고 현황 파일 업데이트"
    msg["From"] = "no1gunhee@gmail.com"
    msg["To"] = email
    # msg["To"] = "eunah7603@naver.com"
    #msg["To"] = "no1gunhee@gmail.com"

    # 메일 내용 쓰기
    content = "날짜 기준 최신 데이터를 전달드립니다.\n\n"
    content_part = MIMEText(content, "plain")
    msg.attach(content_part)

    # 데이터 파일 첨부하기
    with open(file_name, 'rb') as excel_file : 
        attachment = MIMEApplication( excel_file.read() )
        #첨부파일의 정보를 헤더로 추가
        attachment.add_header('Content-Disposition','attachment', filename=file_name[13:]) 
        msg.attach(attachment)


    # 메일 보내고 서버 끄기
    to_mail = email
    #to_mail = "no1gunhee@gmail.com"
    smpt.sendmail(my_id, to_mail, msg.as_string())  
    smpt.quit()


def show_xlsx(text):
    current_directory = os.getcwd()

    # result_xlsx 디렉토리에 있는 모든 파일 목록을 가져옵니다.
    files = os.listdir("./result_xlsx")

    # 특정 문자열이 포함된 .xlsx 파일들만 필터링해서 출력합니다.
    search_string = "★"
    xlsx_files_with_string = [file for file in files if file.endswith(".xlsx") and search_string in file]
    option = st.selectbox(
        text,
        xlsx_files_with_string,
        index=None,
        format_func=lambda x: x if x else "Select an .xlsx file...",
    )
    return option

current_year = datetime.datetime.now().year
years_list = [str(year) for year in range(current_year, 2003, -1)]

month_list = ['01','02', '03', '04', '05', '05','06', '07', '08','09', '10','11','12']

def tab1_content():
    st.markdown("""---""")
    file_name = ""
    
    st.write("1. 다운로드 후 가공할 영역 선택")
    MIN_MAX_RANGE = (datetime.datetime(2004,3,2), datetime.datetime(datetime.datetime.now().year,datetime.datetime.now().month,datetime.datetime.now().day))
    PRE_SELECTED_DATES = (datetime.datetime(2020,1,1), datetime.datetime(datetime.datetime.now().year,datetime.datetime.now().month,datetime.datetime.now().day))

    selected_min, selected_max = st.slider(
        "다운로드할 데이터의 기간 선택",
        value=PRE_SELECTED_DATES,
        min_value=MIN_MAX_RANGE[0],
        max_value=MIN_MAX_RANGE[1],
    )


    col1, col2, col3, col4 = st.columns([1,1,1,1])
    page2_checkbox = col1.checkbox('품목수 월별',value=True)
    page3_checkbox = col2.checkbox('카테고리별',value=True)
    page4_checkbox = col3.checkbox('원료별',value=True)

    detail_value = st.selectbox( 
        '몇년도 품목수/카테고리/원료를 엑셀에 표시할까요. (다운로드 하는 년도 내의 값을 입력해주세요)',
        (years_list))
    

    date_temp1 = selected_min.strftime("%y%m%d")
    date_temp2 = selected_max.strftime("%y%m%d")

    if col4.button("식품 목록 Download"):
        st.caption("약 35,000개의 데이터를 다운받고 있습니다. 5분 정도의 시간이 소요됩니다.")
        load_wb = load_workbook("./form_xlsx/건강기능식품 품목신고_현황_양식_new.xlsx", data_only=True)
        load_ws = load_wb['Sheet1']
        load_while(load_ws, selected_min, selected_max)
        now = datetime.datetime.now()
        d = now.strftime("%y%m%d")
        file_name = "./result_xlsx/★"+ date_temp1+"~"+ date_temp2 +".xlsx"
        load_wb.save(file_name)  
        load_wb = load_workbook(file_name, data_only=True)
        df = pd.read_excel(file_name)
        
        df['등록일'] = df['등록일'].apply(lambda x: str(x).replace('.0', ''))
        df['등록일2'] = df['등록일'].apply(lambda x: pd.to_datetime(str(x), format='%Y%m%d'))
        df["year"] = df["등록일2"].dt.year
        df["month"] = df["등록일2"].dt.month
        df["day"] = df["등록일2"].dt.day


        df_until_now = df[(df['year']== detail_value)] 
        df_until_now = df_until_now.reset_index()

        if page2_checkbox:
            progress_text = "품목수 월별 in progress. Please wait."
            my_bar2 = st.progress(0, text=progress_text)
            my_bar2.progress(10 , text=progress_text) 
            load_ws2 = load_wb["2023년 품목수_월별"]
            page2(df_until_now, file_name, load_ws2)
            my_bar2.progress(100 , text="품목수 월별 process finish.") 


        if page3_checkbox:
            progress_text = "카테고리별 progress. Please wait."
            my_bar3 = st.progress(0, text=progress_text)
            my_bar3.progress(10 , text=progress_text) 
            load_ws3 = load_wb["2023년 카테고리별"]
            page3_loaded = page3(df_until_now, file_name, load_ws3)
            my_bar3.progress(100 , text="카테고리별 process finish.") 

            
        if page4_checkbox:
            progress_text = "원료별 progress. Please wait."
            my_bar4 = st.progress(0, text=progress_text)
            my_bar4.progress(10 , text=progress_text) 
            load_ws4 = load_wb["2023 원료별"]
            page4_loaded = page4(df_until_now, file_name, load_ws4)
            my_bar4.progress(100 , text="원료별 process finish.") 

        load_wb.save(file_name)

   
    st.markdown("""---""")
    st.write("2. 다운로드한 데이터 확인")
    option = show_xlsx("확인을 위한 데이터 선택")

    if(option == None):
        pass
    else:
        file_name = "./result_xlsx/"+option

        tab1, tab2, tab3, tab4 = st.tabs(["전체", "품목수 월별", "카테고리별", "원료별"])
        df_tab1 = pd.read_excel(file_name, sheet_name="Sheet1")
        df_tab2 = pd.read_excel(file_name, sheet_name="2023년 품목수_월별")
        df_tab3 = pd.read_excel(file_name, sheet_name="2023년 카테고리별")
        df_tab4 = pd.read_excel(file_name, sheet_name="2023 원료별")

        
        with tab1:
            st.dataframe(df_tab1)
        with tab2:
            st.dataframe(df_tab2)
        with tab3:
            st.dataframe(df_tab3)
        with tab4:
            st.dataframe(df_tab4)


    st.markdown("""---""")
    st.write("3. 데이터 다운로드")

    method = st.radio("다운로드 방법을 선택해주세요.", ["PC에 직접 다운로드", "Email로"])

    option2 = show_xlsx("다운로드할 파일을 선택해주세요.")
    if(option2 == None):
        pass
    else:

        xlsx_file_path = "./result_xlsx/" + option2
        zip_file_path = "./result_zip/" + option2 + ".zip"
    

        if method == "PC에 직접 다운로드":
            with zipfile.ZipFile(zip_file_path, "w") as zipf:
                zipf.write(xlsx_file_path, option2)

            with open(zip_file_path, 'rb') as f:
                st.download_button('Download Zip', f, file_name=option2+".zip")  # Defaults to 'application/octet-stream'


        else:
            col1, col2 = st.columns([3,1])
            email_address = col1.text_input("Email Address")
            col2.write("")
            col2.write("")
            if col2.button("Send"):
                if not re.match(r"[^@]+@[^@]+\.[^@]+", email_address):
                    col1.error("유효한 이메일 주소를 입력해주세요.")
                else:
                    with zipfile.ZipFile(zip_file_path, "w") as zipf:
                        zipf.write(xlsx_file_path, option2)

                    my_bar_email_send = st.progress(20, text="Sending email")
                    send(zip_file_path, email_address)
                    my_bar_email_send.progress(100, text="Sending email process finish")




